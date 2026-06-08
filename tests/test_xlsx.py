from datetime import datetime, timezone

from openpyxl import load_workbook

from intercom_summary.export.transcript import conversation_to_markdown
from intercom_summary.export.xlsx import export_xlsx
from intercom_summary.intercom.models import Admin, Contact, Conversation, Message


def _sample():
    return Conversation(
        id="42",
        created_at=datetime(2026, 5, 1, tzinfo=timezone.utc),
        updated_at=datetime(2026, 5, 1, tzinfo=timezone.utc),
        state="closed",
        subject="Login issue",
        assignee=Admin(id="1", name="Ada", email="ada@co.com"),
        contact=Contact(name="Cara", email="cara@x.com"),
        messages=[
            Message(0, "user", "Cara", datetime(2026, 5, 1, tzinfo=timezone.utc), "I can't log in"),
            Message(1, "admin", "Ada", datetime(2026, 5, 1, tzinfo=timezone.utc), "Try resetting"),
        ],
        tags=["login"],
        csat_rating=5,
    )


def test_export_xlsx_has_two_sheets_and_rows(tmp_path):
    out = export_xlsx([_sample()], tmp_path / "x.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary", "Messages"]
    # header + 1 conversation
    assert wb["Summary"].max_row == 2
    # header + 2 messages
    assert wb["Messages"].max_row == 3
    assert wb["Summary"]["A2"].value == "42"


def test_transcript_markdown_contains_both_roles():
    md = conversation_to_markdown(_sample())
    assert "Conversation 42" in md
    assert "Agent" in md and "Customer" in md
    assert "Try resetting" in md
