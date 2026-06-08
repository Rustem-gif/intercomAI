"""Export normalised conversations to a readable XLSX workbook.

Two sheets:
  • Summary  — one row per conversation (scan-friendly overview)
  • Messages — one row per message (filterable full thread)
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from intercom_summary.intercom.models import Conversation

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_SUMMARY_COLS = [
    "Conversation ID", "Subject", "Agent", "Customer", "Customer Email",
    "State", "Created (UTC)", "Updated (UTC)", "Messages",
    "CSAT", "CSAT Remark", "First Response (s)", "Tags", "Link",
]
_MESSAGE_COLS = [
    "Conversation ID", "Seq", "Author Type", "Author", "Timestamp (UTC)", "Part", "Text",
]


def _style_header(ws: Worksheet, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autosize(ws: Worksheet, max_width: int = 80) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(12, longest + 2), max_width)


def _iso(dt) -> str:
    return dt.isoformat() if dt else ""


def build_workbook(conversations: Iterable[Conversation]) -> Workbook:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(_SUMMARY_COLS)

    messages = wb.create_sheet("Messages")
    messages.append(_MESSAGE_COLS)

    for convo in conversations:
        summary.append([
            convo.id,
            convo.subject,
            convo.assignee_name,
            convo.contact.name,
            convo.contact.email,
            convo.state,
            _iso(convo.created_at),
            _iso(convo.updated_at),
            convo.message_count,
            convo.csat_rating if convo.csat_rating is not None else "",
            convo.csat_remark,
            convo.first_response_time if convo.first_response_time is not None else "",
            ", ".join(convo.tags),
            convo.web_url(),
        ])
        for m in convo.messages:
            messages.append([
                convo.id, m.seq, m.author_type, m.author_name,
                _iso(m.created_at), m.part_type, m.text,
            ])

    _style_header(summary, len(_SUMMARY_COLS))
    _style_header(messages, len(_MESSAGE_COLS))
    _autosize(summary)
    _autosize(messages)
    # Wrap long message text.
    text_col = _MESSAGE_COLS.index("Text") + 1
    messages.column_dimensions[get_column_letter(text_col)].width = 100
    for row in messages.iter_rows(min_row=2, min_col=text_col, max_col=text_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return wb


def export_xlsx(conversations: list[Conversation], out_path: str | Path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(conversations)
    wb.save(out)
    return out
