"""One-off client deliverable: every Intercom chat in a date range, as readable transcripts.

A client asked for "all the chats from the last 3 months". That is a hand-off, not a product
feature, so this lives as a standalone script rather than a subcommand: it drives the repo's
Intercom client directly and **never touches the QA database**. Going through
`service.fetch_and_store` would be wrong twice over — the Trash blacklist in
`ConversationsStore.save()` silently drops previously-deleted ids, and 27k client-export rows
would pollute the grading tables.

Output is one ZIP per brand (`intercom/brands.py` maps "Betncare" -> "King Billy"), each
holding an `index.xlsx` plus one Markdown transcript per conversation, foldered by month and
agent.

Four things worth knowing before changing this:

1. **The search is date-only, not per-agent.** `build_search_query([], since, until)` omits the
   `admin_assignee_id` clause entirely, which returns every conversation in the window —
   including the ~24% that have no human assignee (Fin-AI-only and unassigned threads). Looping
   over `list_admins()` instead would be slower *and* silently lose those; `--per-agent` keeps
   that path available only as a fallback if the workspace ever rejects the broad query.
2. **Fetching happens in chunks against a resumable cache.** A 3-month window is ~27k
   conversations and ~27k full-thread GETs, so a single `asyncio.gather` over all of them (what
   `fetch_conversations_for_agents` does) means 27k live coroutines and total loss on a crash at
   90%. Raw payloads are cached to `raw/` as they land; `--resume` picks up where it stopped,
   and `--only-build` re-cuts the ZIPs from that cache without touching the network.
3. **Chats only — tickets are excluded.** Intercom's conversation search returns tickets
   alongside chats (same id namespace, same `"type": "conversation"`), and the client asked for
   chats only. `is_ticket` filters them out of the stub sweep, so a ticket never costs a
   full-thread GET, and again at build time so a cache captured before this rule is re-cut
   correctly. `--include-tickets` restores the old behaviour.
4. **The search is swept in parallel sub-windows.** Intercom's search pages get slower the
   deeper you page, so one sweep of a 3-month range is ~180 pages at roughly ten seconds each —
   half an hour before a single thread is fetched. Chopping the range into overlapping weekly
   windows searched concurrently keeps every sweep shallow. Verified against `total_count`:
   the union is exact, with no boundary gaps.

Usage:
    python scripts/export_client_archive.py --dry-run          # count matches, fetch nothing
    python scripts/export_client_archive.py                    # last 92 days, both phases
    python scripts/export_client_archive.py --since 2026-05-28 --until 2026-08-27
    python scripts/export_client_archive.py --resume           # continue an interrupted run
    python scripts/export_client_archive.py --only-build       # rebuild ZIPs from raw/, offline
"""
from __future__ import annotations

import argparse
import asyncio
import gzip
import json
import re
import time
import unicodedata
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Iterator

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from intercom_summary.intercom.brands import brand_label
from intercom_summary.intercom.client import IntercomClient
from intercom_summary.intercom.fetch import (
    _to_unix,
    build_search_query,
    contact_from_payload,
    is_ticket,
    normalise_conversation,
)
from intercom_summary.intercom.htmltext import html_to_text
from intercom_summary.intercom.models import Admin, Conversation, fmt_duration
from intercom_summary.logging_setup import get_logger
from intercom_summary.settings import settings

log = get_logger("export_client_archive")

# Full threads are fetched in chunks so progress is checkpointed to disk regularly. 200 keeps
# the per-chunk gather small while still amortising the file appends.
CHUNK_SIZE = 200
DEFAULT_DAYS = 92

_EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
_UNSAFE_RE = re.compile(r'[/\\:*?"<>|\x00-\x1f]+')
_REDACTED = "[email hidden]"

_INDEX_COLS = [
    ("Conversation ID", 20), ("Date (UTC)", 20), ("Subject", 60), ("Agent", 22),
    ("Customer", 22), ("Customer Email", 28), ("Brand", 14), ("State", 10),
    ("Messages", 10), ("CSAT", 8), ("CSAT Remark", 30), ("First Response", 16),
    ("Tags", 30), ("Transcript File", 52), ("Intercom Link", 60),
]
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


# ── small helpers ─────────────────────────────────────────────────────────────────
def _slug(value: str, fallback: str) -> str:
    """Filesystem-safe folder/file segment. Keeps Cyrillic — agent names are often Russian."""
    v = unicodedata.normalize("NFKC", (value or "").strip())
    v = _UNSAFE_RE.sub("-", v)
    v = re.sub(r"\s+", " ", v).strip(" .-")
    return v[:60] or fallback


def _redact(text: str) -> str:
    return _EMAIL_RE.sub(_REDACTED, text or "")


def _cell(value: Any) -> Any:
    """Excel rejects control characters and caps cells at 32,767 chars."""
    if not isinstance(value, str):
        return value
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    return value[:32_000]


def _iso(dt: datetime | None) -> str:
    return dt.isoformat() if dt else ""


def _default_window(since: str | None, until: str | None) -> tuple[str, str]:
    now = datetime.now(timezone.utc)
    s = since or (now - timedelta(days=DEFAULT_DAYS)).strftime("%Y-%m-%d")
    u = until or now.strftime("%Y-%m-%d")
    return s, u


# ── raw cache ─────────────────────────────────────────────────────────────────────
class RawCache:
    """Append-only cache of raw conversation payloads, so a run can be resumed or re-cut.

    `fetched.txt` is appended *after* the payloads reach the gzip, so a crash in between costs
    a re-fetch, never a missing conversation. Duplicates are deduped on read.
    """

    def __init__(self, root: Path) -> None:
        self.root = root
        self.root.mkdir(parents=True, exist_ok=True)
        self.payloads = root / "conversations.jsonl.gz"
        self.stubs = root / "stub_ids.txt"
        self.fetched = root / "fetched.txt"
        self.failed = root / "failed.txt"
        self.admins = root / "admins.json"
        self.meta = root / "meta.json"

    @staticmethod
    def _read_ids(path: Path) -> list[str]:
        if not path.exists():
            return []
        return [ln.strip() for ln in path.read_text(encoding="utf-8").splitlines() if ln.strip()]

    def stub_ids(self) -> list[str]:
        return self._read_ids(self.stubs)

    def fetched_ids(self) -> set[str]:
        return set(self._read_ids(self.fetched))

    def failed_ids(self) -> set[str]:
        return set(self._read_ids(self.failed))

    def save_stub_ids(self, ids: list[str]) -> None:
        self.stubs.write_text("\n".join(ids) + "\n", encoding="utf-8")

    def save_admins(self, admins: list[dict]) -> None:
        self.admins.write_text(json.dumps(admins, ensure_ascii=False, indent=1), encoding="utf-8")

    def load_admins(self) -> dict[str, Admin]:
        """id -> Admin, so the build phase can attribute team-assigned conversations offline."""
        if not self.admins.exists():
            return {}
        return {
            str(a["id"]): Admin(id=str(a["id"]), name=a.get("name", ""), email=a.get("email", ""))
            for a in json.loads(self.admins.read_text(encoding="utf-8")) if a.get("id")
        }

    def append(self, payloads: list[dict], ok_ids: list[str], bad_ids: list[str]) -> None:
        if payloads:
            # gzip members concatenate; gzip.open(..., "rt") reads them all back transparently.
            with gzip.open(self.payloads, "at", encoding="utf-8") as fh:
                for p in payloads:
                    fh.write(json.dumps(p, ensure_ascii=False) + "\n")
        if ok_ids:
            with self.fetched.open("a", encoding="utf-8") as fh:
                fh.write("\n".join(ok_ids) + "\n")
        if bad_ids:
            with self.failed.open("a", encoding="utf-8") as fh:
                fh.write("\n".join(bad_ids) + "\n")

    def clear_failed(self) -> None:
        self.failed.unlink(missing_ok=True)

    def iter_payloads(self) -> Iterator[dict]:
        """Yield each cached payload once, in cache order."""
        if not self.payloads.exists():
            return
        seen: set[str] = set()
        with gzip.open(self.payloads, "rt", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    raw = json.loads(line)
                except json.JSONDecodeError:
                    log.warning("Skipping a corrupt cache line.")
                    continue
                cid = str(raw.get("id", ""))
                if not cid or cid in seen:
                    continue
                seen.add(cid)
                yield raw


# ── phase A: fetch ────────────────────────────────────────────────────────────────
def _split_window(since: str, until: str, days: int) -> list[tuple[datetime, datetime]]:
    """Chop the range into sub-windows that can be searched in parallel.

    Intercom's search only offers strict `>` / `<` on `created_at`, so adjacent windows are
    overlapped by a second rather than butted together — a conversation created exactly on a
    boundary would otherwise fall through the crack. The overlap double-counts a handful of
    ids, which the caller's de-duplication absorbs.
    """
    start, end = _to_unix(since), _to_unix(until)
    if start is None or end is None or end <= start:
        return [(datetime.fromtimestamp(start or 0, timezone.utc),
                 datetime.fromtimestamp(end or 0, timezone.utc))]
    step = days * 86_400
    out: list[tuple[datetime, datetime]] = []
    lo = start
    while lo < end:
        hi = min(lo + step, end)
        out.append((datetime.fromtimestamp(lo - 1, timezone.utc),
                    datetime.fromtimestamp(hi + 1, timezone.utc)))
        lo = hi
    return out


def _log_tickets(n: int) -> None:
    if n:
        log.info("Excluded %d ticket(s) from the sweep — this archive is chats only "
                 "(pass --include-tickets to keep them).", n)


async def _collect_stub_ids(
    client: IntercomClient, since: str, until: str, per_agent: bool, limit: int | None,
    window_days: int = 7, sweeps: int = 5, include_tickets: bool = False,
) -> list[str]:
    """Every *chat* id in the window, de-duplicated.

    Tickets are dropped here rather than after fetching, so each one saved is a full-thread
    GET not made — on a 3-month window that is thousands of requests.

    Intercom's search pages get slower the deeper you go — a single sweep of a 3-month range is
    ~180 pages at roughly ten seconds each, i.e. half an hour before the first thread is even
    fetched. Splitting the range into short sub-windows searched concurrently keeps every sweep
    shallow and cuts that to a few minutes. `--limit` keeps the simple sequential path so smoke
    tests still stop early instead of sweeping the whole range first.
    """
    seen: set[str] = set()
    ordered: list[str] = []
    tickets = 0

    def keep(stub: dict[str, Any]) -> str | None:
        """The stub's id if it is a chat we haven't seen, else None."""
        nonlocal tickets
        cid = str(stub.get("id", ""))
        if not cid or cid in seen:
            return None
        if not include_tickets and is_ticket(stub):
            tickets += 1
            return None
        seen.add(cid)
        return cid

    if per_agent:
        # Fallback path only. Note it cannot see unassigned/Fin-only conversations.
        admins = (await client.list_admins()) or []
        queries = [build_search_query([str(a["id"])], since, until) for a in admins if a.get("id")]
        log.info("Per-agent mode: sweeping %d teammate(s).", len(queries))
    elif limit:
        queries = [build_search_query([], since, until)]
    else:
        windows = _split_window(since, until, window_days)
        queries = [build_search_query([], lo, hi) for lo, hi in windows]
        log.info("Sweeping %d sub-window(s) of ~%d day(s), %d at a time…",
                 len(queries), window_days, sweeps)

    if limit or len(queries) == 1:
        for q in queries:
            async for stub in client.search_conversations(q):
                if (cid := keep(stub)) is not None:
                    ordered.append(cid)
                    # `limit` counts chats, so a smoke test still gets `limit` transcripts
                    # however many tickets sit in the same window.
                    if limit and len(ordered) >= limit:
                        _log_tickets(tickets)
                        return ordered
        _log_tickets(tickets)
        return ordered

    sem = asyncio.Semaphore(sweeps)
    done = 0

    async def sweep(query: dict[str, Any]) -> None:
        nonlocal done
        async with sem:
            async for stub in client.search_conversations(query):
                # No await between the check and the insert, so this stays race-free.
                if (cid := keep(stub)) is not None:
                    ordered.append(cid)
        done += 1
        log.info("  swept %d/%d window(s) · %d chat(s) so far",
                 done, len(queries), len(ordered))

    await asyncio.gather(*[sweep(q) for q in queries])
    _log_tickets(tickets)
    return ordered


async def _fetch_chunk(
    client: IntercomClient, ids: list[str], sem: asyncio.Semaphore
) -> tuple[list[dict], list[str]]:
    """Fetch one chunk of full threads. Errors are collected, never raised.

    `IntercomClient._request` gives up after 5 consecutive 429/5xx and raises; at 27k requests
    that will happen eventually. Swallowing it per conversation means one bad patch costs a
    handful of ids (retried by `--resume`) instead of the whole run.
    """
    payloads: list[dict] = []
    failed: list[str] = []

    async def one(cid: str) -> None:
        async with sem:
            try:
                payloads.append(await client.get_conversation(cid))
            except Exception as exc:  # noqa: BLE001 - one bad id must not kill a 27k-run
                log.warning("Failed to fetch %s: %s", cid, str(exc)[:160])
                failed.append(cid)

    await asyncio.gather(*[one(cid) for cid in ids])
    return payloads, failed


async def phase_fetch(
    cache: RawCache, since: str, until: str, concurrency: int,
    per_agent: bool, limit: int | None, resume: bool, include_tickets: bool = False,
) -> None:
    settings.require_intercom()
    client = IntercomClient()
    try:
        # This workspace leaves `assignee` null and names the teammate in `admin_assignee_id`
        # on virtually every payload, so without this roster `normalise_conversation` reports
        # 90%+ of conversations as unassigned. Cached to disk so `--only-build` stays offline.
        roster = await client.list_admins()
        cache.save_admins(roster)
        log.info("Cached %d teammate(s) for agent attribution.", len(roster))

        stub_ids = cache.stub_ids() if resume else []
        if stub_ids:
            log.info("Reusing %d cached stub id(s) — skipping the search sweep.", len(stub_ids))
        else:
            log.info("Searching conversations created %s … %s", since, until)
            stub_ids = await _collect_stub_ids(
                client, since, until, per_agent, limit, include_tickets=include_tickets
            )
            cache.save_stub_ids(stub_ids)
        total = len(stub_ids)
        log.info("%d %s in the window.", total,
                 "conversation(s)" if include_tickets else "chat(s)")

        done = cache.fetched_ids() if resume else set()
        # A resumed run retries whatever failed last time, so the failure list is rebuilt.
        cache.clear_failed()
        todo = [cid for cid in stub_ids if cid not in done]
        if not todo:
            log.info("Nothing left to fetch — cache already holds all %d.", total)
            return
        log.info("Fetching %d full thread(s) (%d already cached, concurrency=%d)…",
                 len(todo), total - len(todo), concurrency)

        cache.meta.write_text(json.dumps(
            {"since": since, "until": until, "total": total,
             "started_at": datetime.now(timezone.utc).isoformat()},
            indent=2), encoding="utf-8")

        sem = asyncio.Semaphore(concurrency)
        started = time.monotonic()
        fetched = 0
        n_failed = 0
        for offset in range(0, len(todo), CHUNK_SIZE):
            chunk = todo[offset:offset + CHUNK_SIZE]
            payloads, failed = await _fetch_chunk(client, chunk, sem)
            cache.append(payloads, [str(p.get("id", "")) for p in payloads], failed)
            fetched += len(payloads)
            n_failed += len(failed)

            elapsed = time.monotonic() - started
            rate = fetched / elapsed if elapsed > 0 else 0
            remaining = len(todo) - fetched - n_failed
            eta = fmt_duration(remaining / rate) if rate > 0 else "?"
            log.info("  %d/%d (%.0f%%) · %s elapsed · ETA %s%s",
                     fetched + n_failed, len(todo),
                     100 * (fetched + n_failed) / len(todo),
                     fmt_duration(elapsed), eta,
                     f" · {n_failed} failed" if n_failed else "")

        log.info("Fetch complete: %d cached, %d failed.", fetched, n_failed)
        if n_failed:
            log.warning("Re-run with --resume to retry the %d failed conversation(s).", n_failed)
    finally:
        await client.aclose()


# ── transcript rendering ──────────────────────────────────────────────────────────
def _attachments_by_seq(raw: dict) -> list[list[dict]]:
    """Attachments per message, in the same order `normalise_conversation` builds messages.

    Attachments are dropped everywhere else in the codebase (`Message` has no field for them),
    so an image-only reply renders as an empty message — not acceptable in a chat archive the
    client reads. The skip rule below deliberately mirrors `intercom/fetch.py` so the lists line
    up index-for-index; `render_markdown` verifies the lengths and degrades gracefully if a
    future change to that function breaks the correspondence.
    """
    out: list[list[dict]] = []
    source = raw.get("source") or {}
    if source.get("body") or source.get("type"):
        out.append(list(source.get("attachments") or []))
    for part in (raw.get("conversation_parts") or {}).get("conversation_parts", []):
        body = html_to_text(part.get("body"))
        ptype = part.get("part_type", "")
        if not body and ptype in ("", "conversation_attribute_updated_by_admin"):
            continue
        out.append(list(part.get("attachments") or []))
    return out


def _customer_identity(convo: Conversation, raw: dict) -> tuple[str, str]:
    """Name and email for the customer.

    The fallback that reads these off the thread's own authors used to live here, because
    `normalise_conversation` left the contact empty. It now lives in `intercom/fetch.py` and
    runs during normalisation, so `convo.contact` is already populated; `raw` is still
    consulted for payloads normalised before that fix.
    """
    if convo.contact.name and convo.contact.email:
        return convo.contact.name, convo.contact.email
    recovered = contact_from_payload(raw)
    return (convo.contact.name or recovered.name, convo.contact.email or recovered.email)


def _role(author_type: str) -> str:
    if author_type == "admin":
        return "🧑‍💼 Agent"
    if author_type in ("user", "contact"):
        return "🙋 Customer"
    return f"⚙️ {author_type or 'system'}"


def render_markdown(
    convo: Conversation, raw: dict, redact: bool, include_system: bool = False
) -> tuple[str, int]:
    """Return the transcript and how many messages it actually shows.

    The count is handed back so the index workbook reports the same number the reader sees;
    `convo.message_count` includes the automation entries filtered out below, and an index that
    says 56 next to a transcript showing 45 just looks broken.
    """
    attachments = _attachments_by_seq(raw)
    if len(attachments) != len(convo.messages):
        attachments = [[] for _ in convo.messages]

    pairs = list(zip(convo.messages, attachments))
    hidden = 0
    if not include_system:
        # Roughly 40% of parts in this workspace are empty bot bookkeeping — quick_reply,
        # language_detection_details, default_assignment, sla_applied — which render as
        # "(no text)" and bury the actual conversation. This is the same rule the grader
        # already uses in `Conversation.transcript_text`: keep every human turn, plus any
        # message that actually carried content.
        kept = [
            (m, a) for m, a in pairs
            if (m.text and m.text.strip()) or a or m.author_type in ("admin", "user", "contact")
        ]
        hidden = len(pairs) - len(kept)
        pairs = kept

    raw_name, raw_email = _customer_identity(convo, raw)
    customer = raw_name or "(unknown)"
    email = (_REDACTED if redact else raw_email) if raw_email else "(none)"

    lines = [
        f"# Conversation {convo.id}",
        "",
        f"- **Brand:** {brand_label(convo.brand)}",
        f"- **Subject:** {convo.display_subject or '(none)'}",
        f"- **Agent:** {convo.assignee_name or '(unassigned)'}",
        f"- **Customer:** {customer} <{email}>",
        f"- **State:** {convo.state}",
        f"- **Created (UTC):** {_iso(convo.created_at) or '?'}",
        f"- **Updated (UTC):** {_iso(convo.updated_at) or '?'}",
        f"- **Messages:** {len(pairs)}",
        f"- **First response:** {fmt_duration(convo.first_response_time)}",
        f"- **Time to close:** {fmt_duration(convo.time_to_close)}",
    ]
    if convo.csat_rating is not None:
        remark = _redact(convo.csat_remark) if redact else convo.csat_remark
        lines.append(f"- **CSAT:** {convo.csat_rating}/5 — {remark or '(no remark)'}")
    if convo.tags:
        lines.append(f"- **Tags:** {', '.join(convo.tags)}")
    if hidden:
        lines.append(f"- **Hidden system events:** {hidden} "
                     "(empty bot/automation entries; use --include-system-events to keep them)")
    lines += [f"- **Open in Intercom:** {convo.web_url()}", "", "---", ""]

    for msg, files in pairs:
        when = _iso(msg.created_at) or "?"
        suffix = f" _({msg.part_type})_" if msg.part_type and msg.part_type != "comment" else ""
        text = _redact(msg.text) if redact else msg.text
        # `_author_name` falls back to the author's email when they have no name, so the byline
        # itself leaks an address unless it goes through the same filter.
        who = _redact(msg.author_name) if redact else msg.author_name
        lines.append(f"**{_role(msg.author_type)} · {who}** · {when}{suffix}")
        lines.append("")
        lines.append(text or "_(no text)_")
        for f in files:
            name = f.get("name") or f.get("content_type") or "attachment"
            url = f.get("url") or ""
            if redact:
                name, url = _redact(name), _redact(url)
            lines.append(f"- 📎 **{name}**" + (f" — {url}" if url else ""))
        lines.append("")
    return "\n".join(lines), len(pairs)


# ── phase B: build ────────────────────────────────────────────────────────────────
class BrandArchive:
    """One ZIP + its index workbook, written as conversations stream past.

    The index uses a write-only workbook: the normal openpyxl mode (`export/xlsx.py`) keeps
    every cell in memory, and the King Billy index alone is ~27k rows x 15 columns.
    """

    def __init__(self, path: Path, label: str, window: str, chats_only: bool = True) -> None:
        self.path = path
        self.label = label
        self.window = window
        self.chats_only = chats_only
        self.count = 0
        self.zf = zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, allowZip64=True)
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet("Index")
        for i, (_, width) in enumerate(_INDEX_COLS, start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = width
        self.ws.freeze_panes = "A2"
        self.ws.auto_filter.ref = f"A1:{get_column_letter(len(_INDEX_COLS))}1"
        header = []
        for name, _ in _INDEX_COLS:
            cell = WriteOnlyCell(self.ws, value=name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            header.append(cell)
        self.ws.append(header)

    def add(self, arcname: str, markdown: str, row: list[Any]) -> None:
        self.zf.writestr(arcname, markdown)
        self.ws.append([_cell(v) for v in row])
        self.count += 1

    def close(self) -> None:
        self.zf.writestr("README.txt", self._readme())
        with TemporaryDirectory() as tmp:
            index = Path(tmp) / "index.xlsx"
            self.wb.save(index)
            self.zf.write(index, "index.xlsx")
        self.zf.close()

    def _readme(self) -> str:
        return "\n".join([
            f"Intercom conversation archive — {self.label}",
            f"Window (conversation created date): {self.window}",
            f"Conversations: {self.count}",
            f"Generated: {datetime.now(timezone.utc).isoformat()} (UTC)",
            "",
            "CONTENTS",
            "  index.xlsx    One row per conversation. The 'Transcript File' column gives the",
            "                path to that conversation's transcript inside this archive.",
            "  YYYY-MM/      Transcripts, foldered by month and then by the agent the",
            "                conversation was assigned to ('_unassigned' when nobody was).",
            "",
            "TRANSCRIPTS",
            "  One Markdown (.md) file per conversation, named <date>_<conversation id>.md.",
            "  Each opens with a header block (brand, subject, agent, customer, state, CSAT,",
            "  tags, response times, a link back to Intercom) followed by the full thread in",
            "  order. Attachments are listed under the message they belong to.",
            "",
            "NOTES",
            ("  Chats only. Intercom tickets are a separate object type and are not included."
             if self.chats_only else
             "  Includes both chats and Intercom tickets."),
            "  All timestamps are UTC.",
            "  'First Response' is Intercom's own first-admin-reply metric, in seconds.",
            "  Internal agent notes are included and labelled, so the record is the full",
            "  thread rather than the customer-visible part only. Empty automation entries",
            "  (quick replies, language detection, auto-assignment) are omitted; each",
            "  transcript states how many were hidden.",
        ])


def phase_build(
    cache: RawCache, out_dir: Path, since: str, until: str,
    split_months: bool, redact: bool, include_system: bool = False,
    include_tickets: bool = False,
) -> dict[str, int]:
    window = f"{since} .. {until}"
    archives: dict[str, BrandArchive] = {}
    admins = cache.load_admins()
    if not admins:
        log.warning("No teammate roster cached (raw/admins.json) — every conversation will be "
                    "filed as unassigned. Re-run the fetch phase to populate it.")

    def archive_for(label: str, month: str) -> BrandArchive:
        key = f"{label}|{month}" if split_months else label
        if key not in archives:
            stem = _slug(label, "unbranded").lower().replace(" ", "-")
            name = f"{stem}_{month}.zip" if split_months else f"{stem}_{since}_{until}.zip"
            archives[key] = BrandArchive(out_dir / name, label,
                                         month if split_months else window,
                                         chats_only=not include_tickets)
            log.info("Opening %s", name)
        return archives[key]

    seen = 0
    tickets = 0
    try:
        for raw in cache.iter_payloads():
            # Re-checked here, not only in the sweep: a cache filled by an earlier run (or by
            # --resume against one) still holds tickets, and --only-build must not put them
            # back into the deliverable.
            if not include_tickets and is_ticket(raw):
                tickets += 1
                continue
            convo = normalise_conversation(raw, known_admins=admins)
            label = brand_label(convo.brand)
            month = convo.created_at.strftime("%Y-%m") if convo.created_at else "unknown-date"
            day = convo.created_at.strftime("%Y-%m-%d") if convo.created_at else "unknown-date"
            agent = _slug(convo.assignee_name, "_unassigned")
            arcname = f"{month}/{agent}/{day}_{convo.id}.md"

            markdown, shown = render_markdown(convo, raw, redact, include_system)
            cust_name, cust_email = _customer_identity(convo, raw)
            email = _REDACTED if (redact and cust_email) else cust_email
            row = [
                convo.id, _iso(convo.created_at), convo.display_subject,
                convo.assignee_name or "(unassigned)", cust_name, email,
                label, convo.state, shown,
                convo.csat_rating if convo.csat_rating is not None else "",
                _redact(convo.csat_remark) if redact else convo.csat_remark,
                convo.first_response_time if convo.first_response_time is not None else "",
                ", ".join(convo.tags), arcname, convo.web_url(),
            ]
            archive_for(label, month).add(arcname, markdown, row)

            seen += 1
            if seen % 2000 == 0:
                log.info("  built %d transcript(s)…", seen)
    finally:
        for arc in archives.values():
            arc.close()

    if tickets:
        log.info("Excluded %d cached ticket(s) from the archive — chats only.", tickets)
    if not seen:
        log.warning("No cached payloads found — run the fetch phase first.")
    return {arc.path.name: arc.count for arc in archives.values()}


# ── entry point ───────────────────────────────────────────────────────────────────
async def _dry_run(
    since: str, until: str, per_agent: bool, out_dir: Path, include_tickets: bool = False
) -> None:
    settings.require_intercom()
    client = IntercomClient()
    try:
        query = build_search_query([], since, until)
        data = await client._request(
            "POST", "/conversations/search",
            json={"query": query, "pagination": {"per_page": 1}},
        )
        total = data.get("total_count")
        log.info("Window %s … %s matches %s conversation(s).", since, until, total)
        # `/conversations/search` counts tickets in that total and offers no way to exclude
        # them (there is no searchable ticket field), so the ticket count comes from the
        # tickets endpoint and the chat estimate is the difference. Without this the dry run
        # promises thousands more transcripts than the real run produces.
        if not include_tickets:
            tdata = await client._request(
                "POST", "/tickets/search",
                json={"query": query, "pagination": {"per_page": 1}},
            )
            tickets = tdata.get("total_count")
            log.info("  of which %s are ticket(s) and will be excluded.", tickets)
            if isinstance(total, int) and isinstance(tickets, int):
                log.info("  ≈ %d chat(s) would be archived.", total - tickets)
        if per_agent:
            admins = await client.list_admins()
            log.info("--per-agent would sweep %d teammate(s) and miss anything unassigned.",
                     len(admins))
        log.info("Would write ZIPs to %s (raw cache in %s).", out_dir, out_dir / "raw")
        log.info("Nothing fetched — drop --dry-run to run for real.")
    finally:
        await client.aclose()


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="export_client_archive",
        description="Export every Intercom conversation in a date range as per-brand ZIPs of "
                    "readable transcripts. Does not touch the QA database.",
    )
    p.add_argument("--since", help="Conversations created after this date (YYYY-MM-DD). "
                                   f"Default: {DEFAULT_DAYS} days ago.")
    p.add_argument("--until", help="Conversations created before this date (YYYY-MM-DD). "
                                   "Default: today.")
    p.add_argument("--out-dir", help="Output directory. Default: "
                                     "EXPORT_DIR/client-archive-<since>_<until>.")
    p.add_argument("--concurrency", type=int, default=8,
                   help="Parallel full-thread fetches (default: 8).")
    p.add_argument("--limit", type=int, help="Cap the number of conversations (smoke testing).")
    p.add_argument("--resume", action="store_true",
                   help="Reuse the raw cache: skip what was fetched, retry what failed.")
    p.add_argument("--only-fetch", action="store_true", help="Fetch into the cache, build nothing.")
    p.add_argument("--only-build", action="store_true",
                   help="Rebuild the ZIPs from the cache. Makes no network calls.")
    p.add_argument("--split-months", action="store_true",
                   help="One ZIP per brand per month instead of one per brand.")
    p.add_argument("--per-agent", action="store_true",
                   help="Fallback: iterate teammates instead of one date-only search. Slower, "
                        "and cannot see unassigned or Fin-AI-only conversations.")
    p.add_argument("--include-tickets", action="store_true",
                   help="Keep Intercom tickets. Off by default: the archive is chats only.")
    p.add_argument("--include-system-events", action="store_true",
                   help="Keep empty bot/automation entries in the transcripts (noisier, but a "
                        "complete record of every conversation part).")
    p.add_argument("--redact-emails", action="store_true",
                   help="Mask email addresses in transcripts and the index.")
    p.add_argument("--dry-run", action="store_true",
                   help="Report the match count and planned outputs, then stop.")
    return p


def main(argv: list[str] | None = None) -> None:
    args = build_parser().parse_args(argv)
    if args.only_fetch and args.only_build:
        build_parser().error("--only-fetch and --only-build are mutually exclusive.")

    since, until = _default_window(args.since, args.until)
    out_dir = (Path(args.out_dir) if args.out_dir
               else settings.export_dir / f"client-archive-{since}_{until}")

    if args.dry_run:
        asyncio.run(_dry_run(since, until, args.per_agent, out_dir, args.include_tickets))
        return

    cache = RawCache(out_dir / "raw")

    if not args.only_build:
        asyncio.run(phase_fetch(
            cache, since, until, args.concurrency, args.per_agent, args.limit, args.resume,
            args.include_tickets,
        ))

    if args.only_fetch:
        log.info("Cache is at %s — re-run with --only-build to produce the ZIPs.", cache.root)
        return

    counts = phase_build(cache, out_dir, since, until, args.split_months,
                         args.redact_emails, args.include_system_events,
                         args.include_tickets)
    if counts:
        log.info("Wrote %d archive(s) to %s:", len(counts), out_dir)
        for name, n in sorted(counts.items()):
            size_mb = (out_dir / name).stat().st_size / 1_048_576
            log.info("  %-46s %6d conversations  %6.1f MB", name, n, size_mb)
        log.info("Total: %d conversation(s).", sum(counts.values()))
    failed = cache.failed_ids()
    if failed:
        log.warning("%d conversation(s) could not be fetched and are NOT in the archive. "
                    "Re-run with --resume to retry them.", len(failed))
    log.info("The raw/ cache is not part of the deliverable — delete it before handing over.")


if __name__ == "__main__":
    main()
